
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.optimize import fsolve
import streamlit as st
from PIL import Image
import math
import WilkeC as wc

# In[2]:
st.set_page_config(layout="wide", page_title="Diffusivity AcA")
st.title("DIFFUSIVITY IN TERNARY SYSTEMS")
st.header("EXTRACTION OF ACETIC ACID FROM WATER")

st.write(r"""The diffusivity calculations for Acetic Acid in Water and organic phase are done using the **MODIFIED WILKE CHANG EQUATION**, where the original equation for pure component diffusivities 
            is modified to accomodate concentration variations for *Ternary* or higher order systems, and the **PERKINS AND GEANKOPLIS EQUATION**, which estimates diffusivity in mixtures 
            using *pure component* diffusivities """)

#st.caption("Poling, B. E., Prausnitz, J. M., & O’connell, J. P. (2001). Properties of gases and liquids. McGraw-Hill Education.")

workbook = load_workbook(filename="CalcData.xlsx")
sheet1=workbook["Data"]


# In[3]:


#Arrays for data sets 
phi=np.zeros(3)
code=[1,3,4]
M_wt=np.zeros(3)
mu=np.zeros(3)
Tc=np.zeros(3)
Pc=np.zeros(3)
omega=np.zeros(3)
rho=np.zeros(3)
A=np.zeros(3)
B=np.zeros(3)


with st.sidebar:
    name=st.selectbox("Choose the ORGANIC ENTRAINER in the ternary system",["NBA","IPA"])
    Texp=st.selectbox("Choose the temperature of system",[298,303,308])
    


if name=="NBA":
    code[0]=1
elif name=="IPA":
    code[0]=2
    

    
for i in range(3):
    if code[i]==4:
        mu_w=[0.889,0.797,0.719]
        
    
for i in range(3):
    k=code[i]
    phi[i]=sheet1.cell(row=k+11,column=5).value
    M_wt[i]=sheet1.cell(row=k+11,column=6).value
    mu[i]=sheet1.cell(row=k+11,column=7).value
    Tc[i]=sheet1.cell(row=k+11,column=8).value
    Pc[i]=sheet1.cell(row=k+11,column=9).value
    omega[i]=sheet1.cell(row=k+11,column=10).value
    rho[i]=sheet1.cell(row=k+11,column=11).value
    A[i]=sheet1.cell(row=k+11,column=12).value
    B[i]=sheet1.cell(row=k+11,column=13).value


# In[4]:


#Temperature variation of viscosity (NBA/IPA,AcA ONLY)
if Texp!=298:
    muT=np.exp(A+B/Texp)*rho*M_wt/1000


# In[5]:


org=code[0]

if org==1:
    sheet2=workbook["NBA"]
elif org==2:
    sheet2=workbook["IPA"]


# In[6]:


if Texp==298:
    mu[2]=mu_w[0]
    mucr=np.log(mu)
elif Texp==303:
    muT[2]=mu_w[1]
    mucr=np.log(muT)
elif Texp==308:
    muT[2]=mu_w[2]
    mucr=np.log(muT)


if Texp==298:
    Da_aqp=wc.get_Da_w(Texp,mu[2],phi[2],M_wt[2])
    Da_orgp=wc.get_Da_org(Texp,mu[0],phi[0],M_wt[0])

else:
    Da_aqp=wc.get_Da_w(Texp,muT[2],phi[2],M_wt[2])
    Da_orgp=wc.get_Da_org(Texp,muT[0],phi[0],M_wt[0])



p=101325 #Pa
Tb=391.2 #K
R=8.3144621 #J/mol
Tca=Tc[1] #K
Pca=57.87 #bar
w=0.447
a=0.45724*((R*Tca)**2)/(Pca*100000)
b=0.0778*R*Tca/(Pca*100000)
Tr=Tb/Tca
alpha=(1+(0.37464+1.54226*w-0.26992*w*w)*(1-(Tr**0.5)))**2


# In[8]:


guess=R*Tb/p
def PengR(v):    
    eqn=p-(R*Tb/(v-b))+(a*alpha/(v*(v+b)+b*(v-b)))    
    return eqn


# In[9]:


Va=fsolve(PengR,guess) #m3/mol



# In[10]:

if Texp==298:
    n=2 
elif Texp==303:
    n=8
elif Texp==308:
    n=14


conc_aq=np.zeros(3)
mumix_aq=np.zeros(5)
phi_mix_aq=np.zeros(5)
Da_aqwc=np.zeros(5)
Da_aqpg=np.zeros(5)
for i in range(5):
    for j in range(3):
        conc_aq[j]=sheet2.cell(row=i+n,column=j+12).value
    mum=mucr*conc_aq
    mumix_aq[i]=np.exp(np.sum(mum))
    phim=conc_aq*phi*M_wt
    phi_mix_aq[i]=phim[0]+phim[2]
    Da_aqwc[i]=(7.4e-8)*np.power(phi_mix_aq[i]/1000,0.5)*Texp/(mumix_aq[i]/1000)/np.power(Va*(1e6),0.6)
    muper=np.exp(mucr)
    Da_aqpg[i]=(conc_aq[0]*Da_orgp*np.power(muper[0],0.8)+conc_aq[2]*Da_aqp*np.power(muper[2],0.8))/np.power(mumix_aq[i],0.8)


# In[11]:


conc_org=np.zeros(3)
mumix_org=np.zeros(5)
phi_mix_org=np.zeros(5)
Da_orgwc=np.zeros(5)
Da_orgpg=np.zeros(5)
for i in range(5):
    for j in range(3):
        conc_org[j]=sheet2.cell(row=i+n,column=j+15).value
    mum=mucr*conc_org
    mumix_org[i]=np.exp(np.sum(mum))
    phim=conc_org*phi*M_wt
    phi_mix_org[i]=phim[0]+phim[2]
    Da_orgwc[i]=(7.4e-8)*np.power(phi_mix_org[i]/1000,0.5)*Texp/(mumix_org[i]/1000)/np.power(1.62*Va*(1e6),0.6)
    muper=np.exp(mucr)
    Da_orgpg[i]=(conc_org[0]*Da_orgp*np.power(muper[0],0.8)+conc_org[2]*Da_aqp*np.power(muper[2],0.8))/np.power(mumix_org[i],0.8)


Da_orgexp=np.zeros(5)
for i in range(5):
    Da_orgexp[i]=sheet2.cell(row=i+n,column=9).value

Da_aqexp=np.zeros(5)
for i in range(5):
    Da_aqexp[i]=sheet2.cell(row=i+n,column=8).value

conc_AcAorg=np.zeros(5)
for i in range(5):
    conc_AcAorg[i]=sheet2.cell(row=i+n,column=16).value

conc_AcAaq=np.zeros(5)
for i in range(5):
    conc_AcAaq[i]=sheet2.cell(row=i+n,column=13).value

error_aqwc=(Da_aqwc-Da_aqexp)/Da_aqwc*100
error_aqpg=(Da_aqpg-Da_aqexp)/Da_aqpg*100
error_orgwc=(Da_orgwc-Da_orgexp)/Da_orgwc*100
error_orgpg=(Da_orgpg-Da_orgexp)/Da_orgpg*100
# rmsd_aq=np.sqrt(np.sum(np.power(error_aq,2))/5)
# rmsd_org=np.sqrt(np.sum(np.power(error_org,2))/5)

fig1,ax1=plt.subplots()
ax1.plot(conc_AcAorg,Da_orgexp*(1e5),'ko'),
ax1.plot(conc_AcAorg,Da_orgwc*(1e5),'k-')
ax1.plot(conc_AcAorg,Da_orgpg*(1e5),'k--')
ax1.legend(['GROMACS', 'Wilke-Chang','Perkins-Geankoplis'], bbox_to_anchor=(1.8, 1), ncol=2)
plt.xlabel("Mole fraction of AcA in organic phase")
plt.ylabel("Diffusivity in orgnic phase ($10^{-5} cm^2 s^{-1}$)")
plt.title('Diffusivity vs mol fraction of acetic acid in organic phase',pad=2)


fig2,ax2=plt.subplots()
ax2.plot(conc_AcAaq,Da_aqexp*(1e5),'ko'),
ax2.plot(conc_AcAaq,Da_aqwc*(1e5),'k-')
ax2.plot(conc_AcAaq,Da_aqpg*(1e5),'k--')
ax2.legend(['GROMACS','Wilke-Chang','Perkins-Geankoplis'], bbox_to_anchor=(1.8, 1), ncol=2)
plt.xlabel("Mole fraction of AcA in aqueous phase")
plt.ylabel("Diffusivity in aqueous phase ($10^{-5} cm^2 s^{-1}$)")
plt.title('Diffusivity vs mol fraction of acetic acid in aqueous phase',pad=2)

st.pyplot(fig1)

with st.expander(label="Raw data of diffusivity in organic phase (error in %)"):
    for i in range(5):
        cols=st.columns(4)
        cols[1].metric('GROMACS',np.around(Da_orgexp[i]*(1e5),decimals=3))
        cols[2].metric('Wilke-Chang',np.around(Da_orgwc[i]*(1e5),decimals=3),np.around(error_orgwc[i],decimals=2))
        cols[3].metric('Perkins-Geankoplis',np.around(Da_orgpg[i]*(1e5),decimals=3),np.around(error_orgpg[i],decimals=2))
        cols[0].metric('AcA in organic phase (mol/mol)',np.around(conc_AcAorg[i],decimals=3))

st.pyplot(fig2)


with st.expander(label="Raw data of diffusivity in aqueous phase (error in %)"):
    for i in range(5):
        cols=st.columns(4)
        cols[1].metric('GROMACS',np.around(Da_aqexp[i]*(1e5),decimals=3))
        cols[2].metric('Wilke-Chang',np.around(Da_aqwc[i]*(1e5),decimals=3),np.around(error_aqwc[i],decimals=2))
        cols[3].metric('Perkins-Geankoplis',np.around(Da_aqpg[i]*(1e5),decimals=3),np.around(error_aqpg[i],decimals=2))
        cols[0].metric('AcA in aqueous phase (mol/mol)',np.around(conc_AcAaq[i],decimals=3))

# st.write("RMSD in aqueous phase", rmsd_aq)
# st.write("RMSD in organic phase", rmsd_org)


# if Texp==298:
#     Da_aqp=wc.get_Da_w(Texp,mu[2],phi[2],M_wt[2])
#     Da_orgp=wc.get_Da_org(Texp,mu[0],phi[0],M_wt[0])

# else:
#     Da_aqp=wc.get_Da_w(Texp,muT[2],phi[2],M_wt[2])
#     Da_orgp=wc.get_Da_org(Texp,muT[0],phi[0],M_wt[0])

# print(Da_aqp)

st.header("Reference")
st.caption("Poling, B. E., Prausnitz, J. M., & O’connell, J. P. (2001). Properties of gases and liquids. McGraw-Hill Education.")

